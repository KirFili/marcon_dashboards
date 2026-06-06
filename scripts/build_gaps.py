"""Рабочий лист пробелов справочника: штук/коробке, коробов/паллету, остаток
в паллетах — ПРЯМЫЕ значения 1С (без фолбэка), по всем позициям отчёта.

Запуск: PYTHONPATH=. uv run python scripts/build_gaps.py
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.inventory import BOX_UNITS, PIECE_UNITS, WEIGHT_UNITS, parse_skus, parse_stock

SAMPLES = Path("data/samples")
OUT = SAMPLES / "05_sku_pallet_gaps_04062026.xlsx"

COLS = [
    "Код 1С", "Название", "Штук в коробке", "Коробок на паллете",
    "Остаток на 04.06.2026 (сырой)", "Ед.", "Остаток в паллетах", "Комментарий",
]
WIDTHS = [14, 48, 14, 16, 22, 8, 16, 46]


def main() -> None:
    refs = parse_skus(SAMPLES / "01_skus.xls")
    stock = parse_stock(SAMPLES / "03_stock_04062026.xlsx")

    rows = []
    for s in stock:
        ref = refs.get(s.code)
        if ref is None:
            rows.append([s.code, s.name, None, None, round(s.closing, 2), "", None,
                         "Код 1С нет в справочнике"])
            continue
        upb = ref.units_per_box       # прямое [124]
        bpp = ref.boxes_per_pallet    # прямое [67]
        unit = ref.unit
        notes, pallets = [], None

        if unit in BOX_UNITS:
            boxes = s.closing
            if bpp:
                pallets = round(boxes / bpp, 2)
            else:
                notes.append("нет «коробок на паллете»")
        elif unit in PIECE_UNITS:
            if not upb:
                notes.append("нет «штук в коробке»")
            if not bpp:
                notes.append("нет «коробок на паллете»")
            if upb and bpp:
                pallets = round((s.closing / upb) / bpp, 2)
        elif unit in WEIGHT_UNITS:
            notes.append(f"весовой/наливной ({ref.unit}) — не в коробках")
        else:
            notes.append(f"ед. «{ref.unit}» не переводится в паллеты")

        rows.append([
            ref.code, ref.name, upb, bpp, round(s.closing, 2), ref.unit, pallets,
            "; ".join(notes),
        ])

    _write(rows)

    n = len(rows)
    no_upb = sum(1 for r in rows if "штук в коробке" in r[7])
    no_bpp = sum(1 for r in rows if "коробок на паллете" in r[7])
    nf = sum(1 for r in rows if "нет в справочнике" in r[7])
    weight = sum(1 for r in rows if "весовой" in r[7])
    has_pallets = sum(1 for r in rows if r[6] is not None)
    print(f"Файл: {OUT}")
    print(f"Строк: {n} | остаток в паллетах посчитан: {has_pallets}")
    print(f"  нет «коробок на паллете»: {no_bpp}")
    print(f"  нет «штук в коробке»: {no_upb}")
    print(f"  весовой/наливной: {weight}")
    print(f"  кода нет в справочнике: {nf}")


def _write(rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пробелы справочника"
    hf = PatternFill("solid", fgColor="1F4E78")
    ff = Font(bold=True, color="FFFFFF")
    th = Side(style="thin", color="D0D0D0")
    bd = Border(th, th, th, th)
    ws.append(COLS)
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(1, ci)
        c.fill, c.font, c.border = hf, ff, bd
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for r in rows:
        ws.append(r)
    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = bd
            c.alignment = Alignment(vertical="top", wrap_text=(c.column == len(COLS)))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT)


if __name__ == "__main__":
    main()
