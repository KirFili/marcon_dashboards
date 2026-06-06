"""Строит таблицу сверки паллет по профильным SKU из выгрузок 1С.

Запуск: uv run python scripts/build_recon.py
"""

from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.inventory import compute_pallets, is_profile, parse_skus, parse_stock

SAMPLES = Path("data/samples")
SKUS = SAMPLES / "01_skus.xls"
STOCK = SAMPLES / "03_stock_04062026.xlsx"
OUT = SAMPLES / "04_recon_pallets_01012025-04062026.xlsx"

COLS = [
    "Код 1С", "Наименование", "Ед. хранения", "Камера",
    "Занято паллетомест на 01.01.2025", "Принято паллет", "Отгружено паллет",
    "Остаток паллетомест на 04.06.2026", "Остаток коробок на 04.06.2026", "Комментарий",
]
WIDTHS = [14, 46, 12, 18, 16, 13, 14, 16, 16, 44]


def main() -> None:
    refs = parse_skus(SKUS)
    stock = parse_stock(STOCK)

    rows, stat = [], Counter()
    for s in stock:
        ref = refs.get(s.code)
        if ref is None:
            stat["нет в справочнике"] += 1
            continue
        if not is_profile(ref):
            stat["непрофильный"] += 1
            continue
        c = compute_pallets(ref, s)
        computed = c.opening_slots is not None and c.closing_slots is not None
        stat["профиль: посчитано" if computed else "профиль: нет данных"] += 1
        rows.append([
            ref.code, ref.name, ref.unit, c.chamber,
            c.opening_slots, c.inbound_pallets, c.outbound_pallets,
            c.closing_slots, c.closing_boxes, c.note,
        ])

    _write(rows)

    total_profile = stat["профиль: посчитано"] + stat["профиль: нет данных"]
    print(f"Файл: {OUT}")
    print(f"Всего строк отчёта: {len(stock)}")
    for k, v in stat.most_common():
        print(f"  {v:4} {k}")
    if total_profile:
        pct = 100 * stat["профиль: посчитано"] // total_profile
        print(f"\nПокрытие профиля: {stat['профиль: посчитано']}/{total_profile} ({pct}%)")

    # спот-чек
    chk = next((r for r in rows if r[0] == "98955"), None)
    if chk:
        print(f"Спот-чек 98955: остаток паллетомест = {chk[7]} (ожидали 32)")


def _write(rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сверка паллет"
    hf = PatternFill("solid", fgColor="1F4E78")
    ff = Font(bold=True, color="FFFFFF")
    th = Side(style="thin", color="D0D0D0")
    bd = Border(th, th, th, th)
    ws.append(COLS)
    for ci in range(1, len(COLS) + 1):
        c = ws.cell(1, ci)
        c.fill, c.font, c.border = hf, ff, bd
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for r in rows:
        ws.append(r)
    for ci, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = bd
            c.alignment = Alignment(vertical="top", wrap_text=(c.column == len(COLS)))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUT)


if __name__ == "__main__":
    main()
